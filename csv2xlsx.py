import sys
import subprocess
import os
import re
from pathlib import Path
import shutil 
# Imports for external libraries (pandas, openpyxl) will be deferred until after installation

# --- Dependency Installation ---
REQUIRED_PACKAGES = ['pandas', 'openpyxl', 'xlsxwriter'] 

def install_packages():
    """Checks for and installs required packages using pip in the active environment."""
    print("Checking for required packages...")
    
    try:
        # Installs all required packages using pip in the active environment (venv)
        subprocess.check_call([sys.executable, "-m", "pip", "install", *REQUIRED_PACKAGES])
        print("Required packages are installed successfully or were already present.")
    except subprocess.CalledProcessError as e:
        print(f"\nERROR: Failed to install required packages: {e}")
        print("Please ensure you are running this script from an active virtual environment (venv) and have internet access.")
        sys.exit(1)
        
# --- Formatting Helper Function ---
def apply_alternating_row_color(sheet, start_row, end_row, end_col_letter):
    """Applies a light gray zebra stripe conditional formatting rule to a sheet."""
    import openpyxl as op
    
    # Define light gray fill color
    gray_fill = op.styles.PatternFill(start_color="EDEDED", end_color="EDEDED", fill_type="solid")
    dxf = op.styles.differential.DifferentialStyle(fill=gray_fill)
    
    # Define the range to apply the formatting rule
    range_str = f"A{start_row}:{end_col_letter}{end_row}"
    
    # Create the rule: MOD(ROW(), 2)=0 highlights even rows (light gray)
    rule = op.formatting.rule.Rule(type="expression", dxf=dxf, stopIfTrue=True)
    rule.formula = ["MOD(ROW(), 2)=0"]
    
    sheet.conditional_formatting.add(range_str, rule)

# --- Configuration ---
PROFILE_TEMPLATE_FILENAME = "template.xlsx" 
OUTPUT_FORMAT = ".xlsx"
# ---------------------

def extract_and_convert(case_notes_path: Path, template_path: Path):
    """
    Reads CSV data, copies template, and writes all data (Profile, Case Notes)
    into the copied template, applying alternating row colors and preserving all formatting.
    """
    # Import libraries here, now guaranteed to be installed.
    import pandas as pd 
    import openpyxl as op 
    
    try:
        print(f"Processing: {case_notes_path.name}")

        # --- 1. Extract Metadata and Case Notes Data from CSV ---
        with case_notes_path.open('r', encoding='utf-8') as f:
            lines = f.readlines()
        
        client_name, hmis_number = "UNKNOWN CLIENT", "N/A"
        data_lines = [line.strip() for line in lines if line.strip() and not line.strip().startswith('#')]
        
        # Extract Client Name (Row 1) and HMIS # (Row 2)
        if len(data_lines) >= 2:
            if ',' in data_lines[0].strip():
                full_client_string = data_lines[0].strip().split(',', 1)[1].strip().strip('"')
                
                # FIX 1: Truncate client name after the name (before the ID/Case Notes)
                # Search for the start of a number or the string " Case Notes" and cut before it.
                match = re.search(r'\s\d+| Case Notes', full_client_string)
                if match:
                    client_name = full_client_string[:match.start()].strip()
                else:
                    client_name = full_client_string.strip()
                    
            if ',' in data_lines[1].strip():
                # HMIS # extracted from the second line
                hmis_number = data_lines[1].strip().split(',', 1)[1].strip().strip('"')

        # Find Case Notes Header
        header_index = -1
        case_notes_header_search = 'Date,Staff,Note'
        for i, line in enumerate(lines):
            if line.strip().startswith(case_notes_header_search):
                header_index = i
                break

        if header_index == -1:
            print(f"Skipping: '{case_notes_path.name}'. Could not find Case Notes header.")
            return

        # Read Case Notes Data
        case_notes_df = pd.read_csv(
            case_notes_path,
            skiprows=header_index,
            header=0,
            encoding='utf-8'
        )
        
        case_notes_df['Date'] = pd.to_datetime(case_notes_df['Date'], errors='coerce')
        case_notes_df = case_notes_df.dropna(subset=['Date'])
        
        # Sort in descending order
        case_notes_df_sorted = case_notes_df.sort_values(by='Date', ascending=False)
        
        if case_notes_df_sorted.empty:
            print(f"Skipping: '{case_notes_path.name}'. Case notes data is empty after sorting.")
            return
            
        # Ensure 'Date' column is date-only for Excel insertion
        case_notes_df_sorted['Date'] = case_notes_df_sorted['Date'].dt.date

        # --- 2. Create Output File by Copying Template ---
        base_name = case_notes_path.name.rsplit('.', 1)[0]
        base_name_cleaned = re.sub(r'_extracted', '', base_name, flags=re.IGNORECASE)
        output_filename = f"{base_name_cleaned}{OUTPUT_FORMAT}"
        output_path = case_notes_path.parent / output_filename
        
        shutil.copy2(template_path, output_path)
        print(f"  --> Copied template to: {output_path.name}")

        # --- 3. Load Copied File and Insert Data ---
        workbook = op.load_workbook(output_path)
        
        # --- Sheet 1: Profile (B1 and B3 ONLY) ---
        profile_sheet = workbook.worksheets[0]
        try:
            profile_sheet = workbook['Profile']
        except KeyError: pass

        # B1: Name (Row 1, Column 2) - Truncated Name
        profile_sheet.cell(row=1, column=2, value=client_name) 
        
        # FIX 2: HMIS # now goes to B3 (Row 3, Column 2)
        profile_sheet.cell(row=3, column=2, value=hmis_number) 
        
        print("  --> Updated Profile sheet (only B1 and B3).")
        
        # --- Sheet 2: Case Notes (Date Descending, No Time) ---
        notes_sheet = workbook.worksheets[1]
        try:
            notes_sheet = workbook['Case Notes']
        except KeyError: pass

        # Clear any existing placeholder data starting from row 2
        notes_sheet.delete_rows(2, notes_sheet.max_row) 

        # Insert sorted data starting from the second row (row=2)
        data_to_write = case_notes_df_sorted.values.tolist()
        for row_index, row_data in enumerate(data_to_write, start=2):
            for col_index, value in enumerate(row_data, start=1):
                notes_sheet.cell(row=row_index, column=col_index, value=value)

        # --- Sheet 3: Room Checks (Blank Placeholder) ---
        room_sheet = None
        try:
            room_sheet = workbook['Room Checks']
            # Clear any data to ensure it is a blank placeholder
            room_sheet.delete_rows(2, room_sheet.max_row)
        except KeyError:
            print("WARNING: Could not find a sheet named 'Room Checks'. Skipping conditional formatting for it.")

        # --- 4. Apply Column Widths and Alternating Row Colors ---
        
        # Apply Column Widths for Case Notes
        notes_sheet.column_dimensions['A'].width = 12.0
        notes_sheet.column_dimensions['B'].width = 10.0
        notes_sheet.column_dimensions['C'].width = 80.0
        
        # Apply Alternating Colors (Zebra Stripes)
        # Profile (A1 to E10 assumed range)
        apply_alternating_row_color(profile_sheet, 1, 10, 'E') 
        
        # Case Notes (Data starts row 2)
        apply_alternating_row_color(notes_sheet, 2, len(case_notes_df_sorted) + 1, 'C')
        
        # Room Checks (Apply to first 100 rows for blank sheet)
        if room_sheet:
             apply_alternating_row_color(room_sheet, 2, 100, 'F') 
        
        print("  --> Applied alternating row colors to all sheets.")
        
        # --- 5. Save the Updated Template Copy ---
        workbook.save(output_path)
        print(f"  --> Successfully saved data to: {output_path.name}")

    except Exception as e:
        print(f"An error occurred while processing {case_notes_path.name}: {e}")

# --- Main Execution Block ---
if __name__ == "__main__":
    
    # 1. Automatically install packages
    install_packages()
    
    # 2. Process files
    current_dir = Path.cwd()
    print(f"\nStarting file search in directory: {current_dir}")

    profile_template_path = current_dir / PROFILE_TEMPLATE_FILENAME
    if not profile_template_path.exists():
        print(f"\nERROR: Template file not found at {profile_template_path}")
        print(f"Please ensure '{PROFILE_TEMPLATE_FILENAME}' is an **XLSX** file in the same directory as this script.")
    else:
        for csv_file in current_dir.rglob('*.csv'):
            if 'template.xlsx' not in csv_file.name and 'Demo Case Notes .xlsx' not in csv_file.name: 
                extract_and_convert(csv_file, profile_template_path)

        print("\nBatch conversion complete.")
